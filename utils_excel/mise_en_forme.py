# Copyright ou © ou Copr. David Sanchez, (11/05/2022)

# david.sanchez@insa-toulouse.fr

# Ce logiciel est un programme informatique servant à créer des feuilles au
# format Excel utilisées pour la correction des copies.

# Ce logiciel est régi par la licence [CeCILL|CeCILL-B|CeCILL-C] soumise au
# droit français et respectant les principes de diffusion des logiciels libres.
# Vous pouvez utiliser, modifier et/ou redistribuer ce programme sous les
# conditions de la licence [CeCILL|CeCILL-B|CeCILL-C] telle que diffusée par le
# CEA, le CNRS et l'INRIA sur le site "http://www.cecill.info".

# En contrepartie de l'accessibilité au code source et des droits de copie,
# de modification et de redistribution accordés par cette licence, il n'est
# offert aux utilisateurs qu'une garantie limitée.  Pour les mêmes raisons,
# seule une responsabilité restreinte pèse sur l'auteur du programme,  le
# titulaire des droits patrimoniaux et les concédants successifs.

# A cet égard  l'attention de l'utilisateur est attirée sur les risques
# associés au chargement,  à l'utilisation,  à la modification et/ou au
# développement et à la reproduction du logiciel par l'utilisateur étant
# donné sa spécificité de logiciel libre, qui peut le rendre complexe à
# manipuler et qui le réserve donc à des développeurs et des professionnels
# avertis possédant  des  connaissances  informatiques approfondies.  Les
# utilisateurs sont donc invités à charger  et  tester  l'adéquation  du
# logiciel à leurs besoins dans des conditions permettant d'assurer la
# sécurité de leurs systèmes et ou de leurs données et, plus généralement,
# à l'utiliser et l'exploiter dans les mêmes conditions de sécurité.

# Le fait que vous puissiez accéder à cet en-tête signifie que vous avez
# pris connaissance de la licence [CeCILL|CeCILL-B|CeCILL-C], et que vous en
# avez accepté les termes.

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Alignment, PatternFill, Side


def ColorFill(couleur):
    return PatternFill(start_color=couleur, end_color=couleur,
                       fill_type='solid')


Rotation = Alignment(text_rotation=90, wrap_text=True)


# Mise en forme d'une ligne
introw = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))
Gaucherow = Border(left=Side(style='medium'),
                   right=Side(style='thin'),
                   top=Side(style='medium'),
                   bottom=Side(style='medium'))
Droiterow = Border(left=Side(style='thin'),
                   right=Side(style='medium'),
                   top=Side(style='medium'),
                   bottom=Side(style='medium'))


def bordure_ligne(ws, Cdeb, Cfin, Ligne, couleur, colorier):
    '''
    Dans la feuille active ws du fichier excel, met en forme une ligne
    encadrée en médium avec des séparations en trait fin
    Si colorier = True, rempli avec couleur si le numéro de la ligne est pair
    '''
    # On met en forme les cases intérieures
    for c in range(Cdeb+1, Cfin):
        cell = ws[get_column_letter(c)+str(Ligne)]
        cell.border = introw
        if colorier and (Ligne % 2 == 0):
            cell.fill = ColorFill(couleur)
    # On met en forme la case de gauche
    cell = ws[get_column_letter(Cdeb)+str(Ligne)]
    cell.border = Gaucherow
    if colorier and (Ligne % 2 == 0):
        cell.fill = ColorFill(couleur)
    # On met en forme la case de droite
    cell = ws[get_column_letter(Cfin)+str(Ligne)]
    cell.border = Droiterow
    if colorier and (Ligne % 2 == 0):
        cell.fill = ColorFill(couleur)


# Mise en forme d'une colonne
intcol = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
Hautcol = Border(left=Side(style='medium'),
                 right=Side(style='medium'),
                 top=Side(style='medium'),
                 bottom=Side(style='thin'))
Bascol = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='thin'),
                bottom=Side(style='medium'))


def bordure_colonne(ws, Colonne, Ldeb, Lfin, couleur, colorier=False):
    '''
    Dans la feuille active ws du fichier excel, met en forme une colonnee
    encadrée en médium avec des séparations en trait fin
    Si colorier = True, rempli avec couleur si le numéro de la ligne est pair
    '''
    # On met en forme les cases intérieures
    for i in range(Ldeb+1, Lfin):
        cell = ws[get_column_letter(Colonne)+str(i)]
        cell.border = intcol
        if colorier and (i % 2 == 0):
            cell.fill = ColorFill(couleur)
    # On met en forme la cases du haut
    cell = ws[get_column_letter(Colonne)+str(Ldeb)]
    cell.border = Hautcol
    if colorier and (Ldeb % 2 == 0):
        cell.fill = ColorFill(couleur)
    # On met en forme la case du bas
    cell = ws[get_column_letter(Colonne)+str(Lfin)]
    cell.border = Bascol
    if colorier and (Lfin % 2 == 0):
        cell.fill = ColorFill(couleur)


# Mise en forme d'un rectangle dont le nombre de lignes et de colonnes est >1
interieur = Border(left=Side(style='thin'),
                   right=Side(style='thin'),
                   top=Side(style='thin'),
                   bottom=Side(style='thin'))
Haut = Border(left=Side(style='thin'),
              right=Side(style='thin'),
              top=Side(style='medium'),
              bottom=Side(style='thin'))
Bas = Border(left=Side(style='thin'),
             right=Side(style='thin'),
             top=Side(style='thin'),
             bottom=Side(style='medium'))
Gauche = Border(left=Side(style='medium'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
Droite = Border(left=Side(style='thin'),
                right=Side(style='medium'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
CoinHG = Border(left=Side(style='medium'),
                right=Side(style='thin'),
                top=Side(style='medium'),
                bottom=Side(style='thin'))
CoinHD = Border(left=Side(style='thin'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='thin'))
CoinBG = Border(left=Side(style='medium'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='medium'))
CoinBD = Border(left=Side(style='thin'),
                right=Side(style='medium'),
                top=Side(style='thin'),
                bottom=Side(style='medium'))


def bordure_rectangle(ws, Cdeb, Cfin, Ldeb, Lfin,
                      couleur, colorier=False):
    '''
    Dans la feuille active ws du fichier excel, met en forme un
    rectangle encadré en médium avec des séparations en trait fin
    Si colorier = True, rempli avec couleur si le numéro de la ligne est pair
    '''
    for c in range(Cdeb+1, Cfin):
        # On met en forme les cases du haut
        cell = ws[get_column_letter(c)+str(Ldeb)]
        cell.border = Haut
        if colorier and (Ldeb % 2 == 0):
            cell.fill = ColorFill(couleur)
        # On met en forme les cases intérieures
        for r in range(Ldeb+1, Lfin):
            cell = ws[get_column_letter(c)+str(r)]
            cell.border = interieur
            if colorier and (r % 2 == 0):
                cell.fill = ColorFill(couleur)
        # On met en forme les cases du bas
        cell = ws[get_column_letter(c)+str(Lfin)]
        cell.border = Bas
        if colorier and (Lfin % 2 == 0):
            cell.fill = ColorFill(couleur)
    # Mise en forme des bords verticaux
    for r in range(Ldeb+1, Lfin):
        # On met en forme le bord gauche
        cell = ws[get_column_letter(Cdeb)+str(r)]
        cell.border = Gauche
        if colorier and (r % 2 == 0):
            cell.fill = ColorFill(couleur)
        # On met en forme le bord droit
        cell = ws[get_column_letter(Cfin)+str(r)]
        cell.border = Droite
        if colorier and (r % 2 == 0):
            cell.fill = ColorFill(couleur)
    # Mise en forme des 4 coins
    # Haut gauche
    cell = ws[get_column_letter(Cdeb)+str(Ldeb)]
    cell.border = CoinHG
    if colorier and (Ldeb % 2 == 0):
        cell.fill = ColorFill(couleur)
    # Haut droit
    cell = ws[get_column_letter(Cfin)+str(Ldeb)]
    cell.border = CoinHD
    if colorier and (Ldeb % 2 == 0):
        cell.fill = ColorFill(couleur)
    # Bas gauche
    cell = ws[get_column_letter(Cdeb)+str(Lfin)]
    cell.border = CoinBG
    if colorier and (Lfin % 2 == 0):
        cell.fill = ColorFill(couleur)
    # Bas droit
    cell = ws[get_column_letter(Cfin)+str(Lfin)]
    cell.border = CoinBD
    if colorier and (Lfin % 2 == 0):
        cell.fill = ColorFill(couleur)


# Mise en forme d'une cellule seule
medium_border = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))


def bordure_cellule(ws, Colonne, Ligne, couleur, colorier=False):
    '''
    Dans la feuille active ws du fichier excel, met en forme une cellule
    seule encadrée en médium
    Si colorier = True, rempli avec couleur si le numéro de la ligne est pair
    '''
    cell = ws[get_column_letter(Colonne)+str(Ligne)]
    cell.border = medium_border
    if colorier and (Ligne % 2 == 0):
        cell.fill = ColorFill(couleur)


def bordure(ws, Cdeb, Cfin, Ldeb, Lfin, colorier=False, couleur="00CCCCFF"):
    '''
    Dans la feuille active ws du fichier excel, met en forme n'importe quel
    type de rectangle, entouré en medium avec des séparations intérieures en
    fin
    Si colorier = True, rempli avec couleur si le numéro de la ligne est pair
    '''
    if Cdeb != Cfin:
        if Ldeb != Lfin:
            bordure_rectangle(ws, Cdeb, Cfin, Ldeb, Lfin, couleur, colorier)
        else:
            bordure_ligne(ws, Cdeb, Cfin, Ldeb, couleur, colorier)
    else:
        if Ldeb != Lfin:
            bordure_colonne(ws, Cdeb, Ldeb, Lfin, couleur, colorier)
        else:
            bordure_cellule(ws, Cdeb, Ldeb, couleur, colorier)
