# Copyright ou © ou Copr. David Sanchez, (11/05/2022)

# david.sanchez@insa-toulouse.fr

# Ce logiciel est un programme informatique servant à créer des feuilles au
# format Excel utilisées pour la correction des copies.

# Ce logiciel est régi par la licence [CeCILL|CeCILL-B|CeCILL-C] soumise au
# droit français et respectant les principes de diffusion des logiciels libres.
# Vous pouvez utiliser, modifier et/ou redistribuer ce programme sous les
# conditions de la licence [CeCILL|CeCILL-B|CeCILL-C] telle que diffusée par le
# CEA, le CNRS et l'INRIA sur le site "http://www.cecill.info".

# En contrepartie de l'accessibilité au code source et des droits de copie,
# de modification et de redistribution accordés par cette licence, il n'est
# offert aux utilisateurs qu'une garantie limitée.  Pour les mêmes raisons,
# seule une responsabilité restreinte pèse sur l'auteur du programme,  le
# titulaire des droits patrimoniaux et les concédants successifs.

# A cet égard  l'attention de l'utilisateur est attirée sur les risques
# associés au chargement,  à l'utilisation,  à la modification et/ou au
# développement et à la reproduction du logiciel par l'utilisateur étant
# donné sa spécificité de logiciel libre, qui peut le rendre complexe à
# manipuler et qui le réserve donc à des développeurs et des professionnels
# avertis possédant  des  connaissances  informatiques approfondies.  Les
# utilisateurs sont donc invités à charger  et  tester  l'adéquation  du
# logiciel à leurs besoins dans des conditions permettant d'assurer la
# sécurité de leurs systèmes et ou de leurs données et, plus généralement,
# à l'utiliser et l'exploiter dans les mêmes conditions de sécurité.

# Le fait que vous puissiez accéder à cet en-tête signifie que vous avez
# pris connaissance de la licence [CeCILL|CeCILL-B|CeCILL-C], et que vous en
# avez accepté les termes.

from openpyxl.utils import get_column_letter


def somme(Cdeb, Cfin, Ligne):
    '''
    Renvoie une chaîne de caractère permettant à excel de
    faire la somme des cellules d'une ligne
    '''
    return (f"=SUM({get_column_letter(Cdeb)}{Ligne}"
            f":{get_column_letter(Cfin)}{Ligne})")


def sommeprod(Cdeb, Cfin, Ligne1, Ligne2, pts_par_questions,
              arrondi=False):
    '''
    Renvoie une chaîne de caractère permettant à excel de faire sommeprod
    des cellules de ligne1 avec celles de ligne2
    et divisant par le nombre de points par questions
    Si arrondi = True, on arrondit au 1/4 de point supérieur
    '''
    # Si Cdeb = Cfin, on n'utilise pas SUMPRODUCT pour éviter des messages
    # d'erreurs dans excel
    if Cdeb == Cfin:
        base_formule = (f"{get_column_letter(Cdeb)}{Ligne1}"
                        f"*{get_column_letter(Cdeb)}{Ligne2}"
                        f"/{pts_par_questions}")
    else:
        base_formule = (f"SUMPRODUCT({get_column_letter(Cdeb)}{Ligne1}"
                        f":{get_column_letter(Cfin)}{Ligne1},"
                        f"{get_column_letter(Cdeb)}{Ligne2}:"
                        f"{get_column_letter(Cfin)}{Ligne2})"
                        f"/{pts_par_questions}")
    if arrondi:
        formule = "=CEILING("+base_formule+"*4;1)/4"
    else:
        formule = "=" + base_formule
    return formule


def moyenne(Colonne, Ldeb, Lfin):
    '''
    Renvoie une chaîne de caractère permettant à excel de faire la moyenne
    des cellules de la colonne
    '''
    return (f"=AVERAGE({get_column_letter(Colonne)}{Ldeb}"
            f":{get_column_letter(Colonne)}{Lfin})")


def ecart_type(Colonne, Ldeb, Lfin):
    '''
    Renvoie une chaîne de caractère permettant à excel de faire
    l'écart-type des cellules de la colonne
    '''
    return (f"=STDEV({get_column_letter(Colonne)}{Ldeb}"
            f":{get_column_letter(Colonne)}{Lfin})")
