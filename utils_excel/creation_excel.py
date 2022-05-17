# Copyright ou © ou Copr. David Sanchez, (11/05/2022)

# david.sanchez@insa-toulouse.fr

# Ce logiciel est un programme informatique servant à créer des feuilles au
# format Excel utilisées pour la correction des copies.

# Ce logiciel est régi par la licence [CeCILL|CeCILL-B|CeCILL-C] soumise au
# droit français et respectant les principes de diffusion des logiciels libres.
# Vous pouvez utiliser, modifier et/ou redistribuer ce programme sous les
# conditions de la licence [CeCILL|CeCILL-B|CeCILL-C] telle que diffusée par le
# CEA, le CNRS et l'INRIA sur le site "http://www.cecill.info".

# En contrepartie de l'accessibilité au code source et des droits de copie,
# de modification et de redistribution accordés par cette licence, il n'est
# offert aux utilisateurs qu'une garantie limitée.  Pour les mêmes raisons,
# seule une responsabilité restreinte pèse sur l'auteur du programme,  le
# titulaire des droits patrimoniaux et les concédants successifs.

# A cet égard  l'attention de l'utilisateur est attirée sur les risques
# associés au chargement,  à l'utilisation,  à la modification et/ou au
# développement et à la reproduction du logiciel par l'utilisateur étant
# donné sa spécificité de logiciel libre, qui peut le rendre complexe à
# manipuler et qui le réserve donc à des développeurs et des professionnels
# avertis possédant  des  connaissances  informatiques approfondies.  Les
# utilisateurs sont donc invités à charger  et  tester  l'adéquation  du
# logiciel à leurs besoins dans des conditions permettant d'assurer la
# sécurité de leurs systèmes et ou de leurs données et, plus généralement,
# à l'utiliser et l'exploiter dans les mêmes conditions de sécurité.

# Le fait que vous puissiez accéder à cet en-tête signifie que vous avez
# pris connaissance de la licence [CeCILL|CeCILL-B|CeCILL-C], et que vous en
# avez accepté les termes.

from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from utils_excel.mise_en_forme import bordure, Rotation
from utils_excel.formules_excel import somme, sommeprod, moyenne, ecart_type
from re import split


@dataclass
class Option_Feuille_de_Note:
    '''
    Classe contenant toutes les options pour créer le fichier excel
    '''
    nom_du_sujet: str
    colonnes_initiales: list
    # Contient une liste contenant les colonnes initiales ["Nom", "Prénom"]
    liste_questions: list
    # Contient une liste indiquant le nombre de questions pour chaque
    # exercice -- sans doute inutile avec les listes
    nb_max_etud: int
    point_par_question: int
    # en général 1 ou 4 selon que l'on note chaque question sur 1 ou sur 4
    Nom_exercices: list  # Liste contenant le nom de chaque exercice
    Liste_exercices: list
    # Liste de liste, chaque entrée d'une sous-liste contient l'intitulé
    # de la question
    Liste_description_exercices: list
    # liste de liste, chaque entrée d'une sous-liste contient la description
    # de la question
    Liste_bareme: list
    # liste de liste, chaque entrée d'une sous-liste contient le nombre
    # de points de la question
    couleur_ligne: str  # On colorie une ligne sur deux, couleur de la ligne
    arrondir: bool  # Arrondir ou non au 1/4 de point supérieur
    stat: bool  # afficher ou non la moyenne et l'écart-type

    def Create_Feuille_de_notes(self):
        wb = Workbook()  # Création du fichier excell
        ws = wb.active  # Feuille active du workbook
        ws.title = "Correction"  # Titre de la feuille
        excel_color = '00' + split(r'\#', self.couleur_ligne)[-1]
        # Conversion de la couleur au format excel
        colonnes_notes = []  # Liste des colonnes où la note apparait.
        # Pour  les deux dernières colonnes de la liste on trouvera les
        # points du sujet en ligne 4
        liste_bilan = []
        # contient les numéros des colonnes où mettre le bilan exercice
        # par exercice

        #########
        # Entêtes du fichier Excel
        #########
        ws['A1'] = self.nom_du_sujet
        ws['A2'] = "Groupe"
        # Mise en place des colonnes initiales
        col = 1
        for i, val in enumerate(self.colonnes_initiales):
            ws[get_column_letter(col)+'4'] = val
            col += 1
        bordure(ws, 1, col-1, 4, 4)
        bordure(ws, 1, col-1, 5, 4+self.nb_max_etud, True, excel_color)
        # Mise en place de la colonne note
        ws[get_column_letter(col)+'4'] = "Note"
        bordure(ws, col, col, 4, 4)
        bordure(ws, col, col, 5, 4+self.nb_max_etud, True, excel_color)
        colonnes_notes.append(col)
        col += 1
        # Sauvegarde du numéro de la première colonne d'exercices
        col_debut = col
        # Mise en place des entêtes des exercices et leur mise en page
        for i, val in enumerate(self.liste_questions):
            ws[get_column_letter(col)+'1'] = self.Nom_exercices[i]
            for j in range(val):
                ws[get_column_letter(col+j)+'2'] = " " \
                    + self.Liste_exercices[i][j]
                ws[get_column_letter(col+j)+'3'].alignment = Rotation
                ws[get_column_letter(col+j)+'3'] = \
                    self.Liste_description_exercices[i][j]
                ws[get_column_letter(col+j)+'4'] = self.Liste_bareme[i][j]
            bordure(ws, col, col+val-1, 1, 3)
            bordure(ws, col, col+val-1, 4, 4)
            bordure(ws, col, col+val-1, 5, 4+self.nb_max_etud,
                    True, excel_color)
            col += val
        # Mise en place de la colonne note
        ws[get_column_letter(col)+'1'] = "Note"
        bordure(ws, col, col, 1, 3)
        bordure(ws, col, col, 4, 4)
        bordure(ws, col, col, 5, 4+self.nb_max_etud, True, excel_color)
        colonnes_notes.append(col)
        # Sauvegarde du numéro de la dernière colonne d'exercices
        col_finale = col-1
        # On décale de trois colonnes et on met le bilan exercice par exercice
        col += 3
        for k, val in enumerate(self.liste_questions):
            ws[get_column_letter(col)+'1'] = self.Nom_exercices[k]
            bordure(ws, col, col, 1, 3)
            bordure(ws, col, col, 4, 4)
            bordure(ws, col, col, 5, 4+self.nb_max_etud, True, excel_color)
            liste_bilan.append(col)
            col += 1
        # Mise en place de la colonne note
        ws[get_column_letter(col)+'1'] = "Note"
        bordure(ws, col, col, 1, 3)
        bordure(ws, col, col, 4, 4)
        bordure(ws, col, col, 5, 4+self.nb_max_etud, True, excel_color)
        colonnes_notes.append(col)
        # Mise en place de toutes les formules
        # Nombre de points total du sujet
        formule = somme(col_debut, col_finale, 4)
        ws[get_column_letter(colonnes_notes[-2])+'4'] = formule
        ws[get_column_letter(colonnes_notes[-1])+'4'] = formule
        # Calcul de la note dans toutes les colonnes de colonnes_notes
        for i in range(self.nb_max_etud):
            formule = sommeprod(col_debut, col_finale, 4, 5+i,
                                self.point_par_question, self.arrondir)
            for col in colonnes_notes:
                ws[f"{get_column_letter(col)}{5+i}"] = formule
        # Résultats par exercice
        col_exo = col_debut
        for j, val in enumerate(self.liste_questions):
            # Nb de points de l'exercice
            ws[get_column_letter(liste_bilan[j])+'4'] = somme(col_exo,
                                                              col_exo+val-1,
                                                              4)
            # Note de l'étudiant à l'exercice
            for i in range(self.nb_max_etud):
                ws[f"{get_column_letter(liste_bilan[j])}{5+i}"] = sommeprod(
                    col_exo, col_exo+val-1, 4, 5+i, self.point_par_question)
            col_exo += val
        # On rajoute Moyenne et Ecart-type éventuellement
        if self.stat:
            ws[get_column_letter(liste_bilan[0]-1)+str(
                5+self.nb_max_etud)] = "Moyenne"
            ws[get_column_letter(liste_bilan[0]-1)+str(
                6+self.nb_max_etud)] = "Ecart-type"
            for i in range(len(self.Liste_exercices)):
                ws[get_column_letter(liste_bilan[i])+str(
                    5+self.nb_max_etud)] = moyenne(liste_bilan[i], 5,
                                                   4+self.nb_max_etud)
                ws[get_column_letter(liste_bilan[i])+str(
                    6+self.nb_max_etud)] = ecart_type(liste_bilan[i], 5,
                                                      4+self.nb_max_etud)
            ws[get_column_letter(colonnes_notes[-1])+str(
                5+self.nb_max_etud)] = moyenne(colonnes_notes[-1], 5,
                                               4+self.nb_max_etud)
            ws[get_column_letter(colonnes_notes[-1])+str(
                6+self.nb_max_etud)] = ecart_type(colonnes_notes[-1], 5,
                                                  4+self.nb_max_etud)
        return wb
